import openpyxl, os, json, collections
from urllib.parse import urlparse
P=os.path.expanduser
XL=P("~/LAWS-microsite/Data/LAWS_Dataset_V1.xlsx")
IMG=P("~/LAWS-microsite/img")

MANUF={"rtx.com","anduril.com","zala-aero.com","stm.com.tr","gdots.com","kongsberg.com","avinc.com",
 "rafael.co.il","rafael-usa.com","rheinmetall.com","wbgroup.pl","mbda-systems.com","hanwha-aerospace.eu",
 "iai.co.il","uvisionuav.com","elbitsystems.com","helsing.ai","knds.com","saab.com","aselsan.com.tr"}
GOV={"dvidshub.net","navair.navy.mil"}
WIKI={"en.wikipedia.org","commons.wikimedia.org"}

def classify(host):
    h=host.replace("www.","")
    if h in GOV or h.endswith(".mil") or h.endswith(".gov"): return "government","include"
    if h in WIKI: return "wikimedia","include"
    if h in MANUF: return "manufacturer","include"
    return "third-party-editorial","hold"

wb=openpyxl.load_workbook(XL, data_only=True)
rows=list(wb["Source_Photos"].iter_rows(values_only=True))
names={}; declared={}
for r in rows[1:]:
    if not r[0]: continue
    names[r[0]]=r[1]
    for slot,u in zip("abc",(r[2],r[3],r[4])):
        if u and str(u).startswith("http"): declared[(r[0],slot)]=str(u).strip()

srow=list(wb["Systems"].iter_rows(values_only=True)); shdr=srow[0]
sysname={r[0]:dict(zip(shdr,r))["System Name"] for r in srow[1:] if r[0]}
ids=sorted(sysname,key=len,reverse=True)

files=[f for f in sorted(os.listdir(IMG)) if not f.startswith((".","_")) and f not in ("manifest.json","credits.json")]
credits={}; tally=collections.Counter()
for f in files:
    stem=f.rsplit(".",1)[0]
    sid=next((s for s in ids if stem==s or stem.startswith(s+"-")),None)
    slot=stem[len(sid)+1:].lower() if sid and len(stem)>len(sid) else ""
    url=declared.get((sid,slot))
    if url:
        host=urlparse(url).netloc.replace("www.","")
        cat,status=classify(host)
    else:
        host=None; cat="unsourced"; status="hold"
    credits[f]={"systemId":sid,"system":sysname.get(sid),"slot":slot or None,
                "sourceUrl":url,"sourceDomain":host,"category":cat,"status":status}
    tally[(cat,status)]+=1

out={"note":"Per-image attribution. status=include -> may be displayed; status=hold -> file present but must not be rendered until rights are cleared.",
     "policy":{"include":["government","wikimedia","manufacturer"],"hold":["third-party-editorial","unsourced"]},
     "images":dict(sorted(credits.items()))}
open(os.path.join(IMG,"credits.json"),"w").write(json.dumps(out,indent=2)+"\n")

print(f"{len(files)} image files classified\n")
for (c,s),n in sorted(tally.items(), key=lambda x:-x[1]): print(f"  {n:4}  {c:22} -> {s}")
inc=sum(n for (c,s),n in tally.items() if s=="include")
print(f"\n  INCLUDE: {inc}   HOLD: {len(files)-inc}")
sysinc=collections.Counter()
for f,m in credits.items():
    if m["status"]=="include": sysinc[m["systemId"]]+=1
print(f"  systems with >=1 displayable photo: {len(sysinc)} / {len(sysname)}")
